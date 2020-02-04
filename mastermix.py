#!/usr/bin/env python
import csv
import sys
import os
from openpyxl import load_workbook, Workbook
from shutil import move
from numpy import arange
import datetime
from time import sleep


assaylist = []
samples = {}


def main():
    original_directory = os.getcwd()
    path = getpath()
    path1 = getplatemap(path)
    platemap1 = platemapsheet1(path1)
    getassays(platemap1)
    samplecount = checksamples(samples)
    mmtemplate = 'mastermix_template.xlsx'
    reagents = open_reagent_list(original_directory)
    MMoutput(mmtemplate, path, reagents, samplecount)
    

def getpath():
    yesno = "x"
    while yesno.lower() not in "yesno":
        print("Attempt automatic folder detection?")
        yesno = input(":")
        if yesno.lower() not in "yesno":
            print("What?")
    if yesno.lower() in "yes":
        # attempts to navigate to most recent directory
        today = datetime.date.today()
        monday_date = today - datetime.timedelta(days=today.weekday())
        week_of = "Week of {}-{}-{}".format(monday_date.strftime("%m"), monday_date.strftime("%d"), monday_date.strftime("%y"))
        if today.weekday() != 0 and today.weekday() != 2:
            previous_date = today - datetime.timedelta(days=1)
        else:
            previous_date = today
        previous_folder = "{}-{}-{}".format(previous_date.strftime("%m"), previous_date.strftime("%d"), previous_date.strftime("%y"))
        try:
            os.chdir('../../Current Year/{}/{}'.format(week_of, previous_folder))
            path = os.getcwd()
        except FileNotFoundError:
            print("Automatic gel map detection failed.")
            path = manual_directory()
    else:
        path = manual_directory()
    return path


def manual_directory():
    # if automated detection fails or isn't wanted
    while True:
            path = input("Enter the path of your platemap file: ")
            if not os.path.isdir(path):
                print("Not a valid directory")
            else:
                break
    return path


def getplatemap(path):
    filelist= []
    for file in os.listdir(path):
        filelist.append(file)
    for file in filelist:
        if 'gel_' in file:
            # platemap should include this str, other files shouldn't
            print('Is this the platemap?')
            print(file)
            yesno = input(':')
            if yesno.lower() in "yes":
                path1 = path + '\\' + file
                return path1
            elif yesno.lower() in 'no':
                # if for some reason another file has the same substr and is detected by mistake
                while True:
                    print('Here is a list of the files in that directory.')
                    print(filelist)
                    try:
                        filenum = int(input('Which would you like to use? (enter a number corresponding to the order of the files):'))
                        platemap = filelist[filenum - 1]
                        print("You selected '{}'. Are you sure?".format(platemap))
                        yesno = input(':')
                        if yesno.lower() in "yes":
                            break
                        elif yesno.lower() not in 'no':
                            print("this is a yes or no question")
                    except TypeError:
                        print("Use a number dummy")
                    except ValueError:
                        print("Use a number dummy")
                    except IndexError:
                        print("Count much? Use a number that refers to one of the files")
                path1 =  path + "\\" + platemap
                return path1        
            else:
                print('You broke it. Start over and answer yes-no questions with yes-no answers')
                sleep(3)
                quit()

def platemapsheet1(path1):
    wb = load_workbook(filename=path1, data_only=True)
    ws = wb.worksheets[0]
    return ws

           
def getassays(ws):
    # I distinguish bottom rows from top rows because they have different rules
    bottom_rows = list(arange(15,431,16))
    # all the rows that will have assay name information
    assay_rows = sorted(list(arange(2,418,16)) + bottom_rows)  
    rowcounter = 0
    columntitles = []
    bottomrowcounter = 0
    # initialized here to avoid recalculation each loop
    botrowlen = len(bottom_rows)
    for row in ws.values:
        # Resets a column value for each new row. The column value allows me to iterate trough each member of the list I made from the row
        # I start at 4 because I know the values I care about will start in column 4
        columncounter = 3
        rowcounter += 1
        if rowcounter in assay_rows:
            columntitles.clear()     
            while True:
                # excel cell format
                cell = '{}{}'.format(chr(64 + columncounter), rowcounter)
                # Will add the value in the current cell to the assay list if it is not already present in the list
                if ws[cell].value not in assaylist and ws[cell].value != None:
                    assaylist.append(ws[cell].value)
                    if ws[cell].value not in samples:
                        samples[(ws[cell].value)] = []
                columntitles.append(ws[cell].value)    
                columncounter += 1
                if columncounter == 15:
                    break
        if (rowcounter - 7) % 16 == 0:
            for a in range(12):
                # each column
                temprowcounter = rowcounter
                for b in range(8):
                    # each row
                    cell = '{}{}'.format(chr(64 + columncounter), temprowcounter)
                    coltitle = columntitles[columncounter - 3]
                    if coltitle != None:
                        if ws[cell].value != ' ' and ws[cell].value != None:
                            samples[coltitle].append(ws[cell].value)
                        if ws[cell].value == 'RNTC_NTC_A_1_1':
                            # signals the end of an assay
                            cell1 = ('{}{}'.format(chr(64 + columncounter), (bottom_rows[bottomrowcounter])))
                            str1 = ws[cell1].value
                            # checks if there are more samples after the end of this assay, suggesting another assay is tucked beneath
                            if str1 != coltitle and str1 != None:
                                columntitles[columncounter - 3] = str1
                                if str1 not in samples:
                                    samples[str1] = [] 
                    temprowcounter += 1
                columncounter += 1
        if rowcounter > bottom_rows[bottomrowcounter]:
                bottomrowcounter += 1
        if rowcounter > bottom_rows[botrowlen - 1]:
            break
        
            
def checksamples(samples):
    badassays = []
    samplecount = {}
    for assays in samples:
        # removing 'assays' that have no samples. This might happen if a note is included in the cells where assays might be
        if len(samples[assays]) == 0:
            badassays.append(assays)
    if len(badassays) > 0:
        print("The following strings were in places where assays should be, but they're probably not assays.")
        print(badassays)
        print('Should they be removed?')
        response = 'a'
        while response.lower() not in 'yesno':
            response = input(':')
        if response.lower() in 'yes':
            for assays in badassays:
                assaylist.remove(assays)
    for assays in assaylist:
        # adding overage to mastermix calculations
        samplecount[assays] = 0
        n = len(samples[assays])
        if n < 5:
            samplecount[assays] = n + 0.5
        elif n / 10 < 4:
            samplecount[assays] = round(n * 1.1)    
        else:
            samplecount[assays] = n + 4
    return samplecount

    
def open_reagent_list(original_directory):
    # references and external dictionary for assay reagent pairings
    os.chdir(original_directory)
    reagents = {}
    with open('assaydictionary.csv', 'r') as a:
        reagent_list = csv.reader(a)
        for rows in reagent_list:
            if rows[0] in assaylist:
                reagents[rows[0]] = rows[1]
    return reagents


def MMoutput(temp, path, reagents, samplecount):
    wb = load_workbook(temp)
    ws = wb.active
    rowcounter = 1
    # this is used to replace outdated information output by our LIMS if we have new primers that aren't present in LIMS
    noadaptorsprimers = [
        'tRNA_Tyr_AA_1'
    ]
    current_assay = 0
    for row in ws.iter_rows():
        columncounter = 1
        if rowcounter % 8 == 0:
            for x in range(3):
                # outputting the assay
                ws.cell(row=rowcounter, column=columncounter).value = assaylist[current_assay]
                # outputting the right reagent
                ws.cell(row=(rowcounter + 2), column=columncounter).value = reagents.get(assaylist[current_assay], "Not Found")
                # note for cases where we have new primers not present in LIMS
                if assaylist[current_assay] in noadaptorsprimers:
                    ws.cell(row=rowcounter, column=columncounter+1).value = 'Use primers without adaptors'
                if assaylist[current_assay] in samples:
                    # sample count
                    ws.cell(row=rowcounter, column=(columncounter + 3)).value = samplecount[assaylist[current_assay]]
                    # primer volume
                    ws.cell(row=(rowcounter + 5), column=(columncounter + 1)).value = 2
                else:
                    # one last saftey net for an assay that wasn't removed in check samples for some reason
                    print("Error: {} slipped through the cracks. Is there something weird about this assay on the platemap?".format(assaylist[current_assay]))
                # reagent volumes, dependent on reagent used
                if reagents.get(assaylist[current_assay]) == 'ZymoTaq':
                    ws.cell(row=(rowcounter + 3), column=columncounter).value = 'DMSO'
                    ws.cell(row=(rowcounter + 3), column=(columncounter + 1)).value = .5
                else:
                    ws.cell(row=(rowcounter + 3), column=columncounter).value = None
                    ws.cell(row=(rowcounter + 3), column=(columncounter + 1)).value = None
                # a special case
                if assaylist[current_assay] == 'ATP7B_112GA_RD_2':
                    ws.cell(row=(rowcounter + 5), column=(columncounter + 1)).value = 3.5
                current_assay += 1
                columncounter += 4
                if current_assay == len(assaylist):
                    break
            if current_assay == len(assaylist):
                break
        if current_assay == len(assaylist):
            break
        rowcounter += 1   
    datestr = str(datetime.date.today()) + ' Secondary PCR Mastermixes.xlsx'
    wb.save(datestr)
    move(datestr, path)
    enda = input("Press enter to end this program")


if __name__ == "__main__":
    main()
